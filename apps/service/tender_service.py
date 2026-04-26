import asyncio
import json
import re
from io import BytesIO
from itertools import combinations
from typing import List

import openpyxl
from fastapi import BackgroundTasks
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import desc, and_, or_

from apps import AppContext
from apps.algorithms.embedding import QwenEmbeddingVectorizer
from apps.document_parser.base import HDocument
from apps.document_parser.markdown_parser import MarkDownParser
from apps.document_parser.pdf_parser import PdfParser
from apps.repository.entity.file_entity import FileRecordEntity
from apps.repository.entity.tender_entity import BidPlagiarismCheckTask, SubBidPlagiarismCheckTask, \
    DocumentSimilarityRecord
from apps.repository.minio_repository import get_file_url, delete_object
from apps.service.milnus_service import create_tender_vector_milvus_db, create_tender_reference_vector_milvus_db, \
    create_tender_topic_vector_milvus_db, create_rm_text_vector_milvus_db, create_main_topic_vector_milvus_db
from apps.service.tender_compliance_service import create_compliance_check_task, parser_tender_topic, parser_document, \
    insert_into_milvus
from apps.web.dto.tender_task import TenderTaskDto, TenderConditionDto, BasePageDto, TenderSimilarityDto
from apps.web.vo.similarity_respose import TenderTaskPage, format_datetime, TenderSimilarityVO, FileRecordVO

from logger_config import get_logger

logger = get_logger(name=__package__)

app_context = AppContext()


class TenderFile:
    """
    标书文件存储类，用于接收orm实体转换的类
    """
    def __init__(self, tender_task_id, file_id, file_name, file_path, file_size, mime_type, business_id):
        self.tender_task_id = tender_task_id
        self.file_id = file_id
        self.file_name = file_name
        self.file_path = file_path
        self.file_size = file_size
        self.mime_type = mime_type
        self.business_id = business_id


def create_plagiarism_check_tasks(tender_task_dto: TenderTaskDto) -> List:
    """
    创建标书查重任务
    :param tender_task_dto: 标书任务求情参数
        :arg task_name: 任务名称
        :arg task_type: 任务类型， 1-查重， 2-合规
        :arg file_ids: 处理标书文件集合（文件id）
    """
    # 根据文件id获取文件管理表中获取对象的信息数据，如文件的类型，文件路径file_path
    with app_context.db_session_factory() as session:
        file_record_list = session.query(FileRecordEntity) \
                     .filter(FileRecordEntity.id.in_(tender_task_dto.file_ids)).all()
        file_name_list = [file_record.file_name for file_record in file_record_list]
        file_id_list = [file_record.id for file_record in file_record_list]
        task = BidPlagiarismCheckTask(
            check_type=tender_task_dto.task_type,
            task_name=tender_task_dto.task_name,
            file_name_list=",".join(file_name_list),
            file_id_list=','.join(map(str, file_id_list)),
            tender_reference_file_id=tender_task_dto.tender_reference_id
        )

        session.add(task)
        session.commit()
        task_id = task.id
        tender_file_list: List[TenderFile] = [TenderFile(
            tender_task_id=task.id,
            file_id=file_record.id,
            file_name=file_record.file_name,
            file_path=file_record.file_path,
            file_size=file_record.file_size,
            mime_type=file_record.mime_type,
            business_id=file_record.business_id
        ) for file_record in file_record_list]
    return tender_file_list, task_id


def start_plagiarism_check(tender_file_list, tender_reference_id):
    """
    开始异步执行查重检测
    :param tasks: 检测任务
    :param background_tasks: 后台任务对象，用于异步执行任务，fastApi自带
    """
    tasks: List = list(combinations(tender_file_list, 2))
    # 异步解析招标文件
    if tender_reference_id:
        handle_tender_tender_reference_file(tender_reference_id)
    for task in tasks:
        # 遍历每个任务，并开始执行检测任务
        plagiarism_check_tasks(task, tender_reference_id)


def plagiarism_check_tasks(task, tender_reference_id = None):
    """
    任务执行
    :param task: 任务数据，并非任务本身，实体数据
    """
    file_record_a, file_record_b = task
    CheckTask(file_record_a, file_record_b, tender_reference_id).execute()


def bid_plagiarism_check(tender_task_dto: TenderTaskDto, background_tasks: BackgroundTasks):
    """
    标书查重
    :param tender_task_dto: 标书任务参数dto
    :param background_tasks: 后台任务对象，用于异步执行任务，fastApi自带
    :return:
    """
    # 解析标书文件
    background_tasks.add_task(tender_file_parser, tender_task_dto.file_ids)
    if tender_task_dto.task_type == 1:
        # 创建标书任务
        tender_file_list, task_id = create_plagiarism_check_tasks(tender_task_dto)
        # 异步重复性处理
        background_tasks.add_task(start_plagiarism_check, tender_file_list, tender_task_dto.tender_reference_id)
    elif tender_task_dto.task_type == 2:
        task_id = create_compliance_check_task(tender_task_dto, background_tasks)
    else:
        pass
    background_tasks.add_task(update_task_process_status, task_id)


async def tender_file_parser(tender_file_ids):
    """
    在标书进行检测前，标书进行解析入库
    :param tender_file_ids:
    :return:
    """
    for tender_file_id in tender_file_ids:
        # 构建解析器
        md_parser = MarkDownParser()
        # 标书转化为图片
        await md_parser.to_images(tender_file_id=tender_file_id)
        # 解析标题
        topics = await parser_tender_topic(tender_file_id)
        # 解析文件内容
        documents = parser_document(tender_file_id)
        insert_into_milvus(tender_file_id, topics, documents)


# 定义后台任务函数
async def tender_file_parser_task(tender_file_ids: List[int]):
    # 这里写入我们之前优化过的多协程逻辑
    MAX_CONCURRENCY = 10
    semaphore = asyncio.Semaphore(MAX_CONCURRENCY)

    async def process_single(file_id):
        async with semaphore:
            md_parser = MarkDownParser()
            await md_parser.to_images(tender_file_id=file_id)
            topics = await parser_tender_topic(file_id)
            documents = parser_document(file_id)
            insert_into_milvus(file_id, topics, documents)

    tasks = [asyncio.create_task(process_single(fid)) for fid in tender_file_ids]
    await asyncio.gather(*tasks)


def update_task_process_status(task_id):
    with app_context.db_session_factory() as session:
        task: BidPlagiarismCheckTask = session.get(BidPlagiarismCheckTask, task_id)
        sub_task_list = session.query(SubBidPlagiarismCheckTask)\
            .filter(SubBidPlagiarismCheckTask.bid_plagiarism_check_task_id == task_id).all()
        task.process_status = "completed"
        for sub_task in sub_task_list:
            if sub_task.process_status == "processing":
                task.process_status = "processing"
                break
        session.add(task)
        session.commit()



def get_tender_task_list(condition: TenderConditionDto):
    """
    :return:
    """
    # created_at_start: str = None
    # created_at_end: str = None
    # 分页逻辑
    page = condition.page_offset  # 当前页码（从 1 开始）
    per_page = condition.page_size  # 每页记录数
    # 计算偏移量
    offset = (page - 1) * per_page
    condition_array = []

    # 筛选任务类型
    if condition.task_type:
        condition_array.append(BidPlagiarismCheckTask.check_type == condition.task_type)

    # 筛选任务状态
    if condition.process_status:
        condition_array.append(BidPlagiarismCheckTask.process_status == condition.process_status)

    # 筛选时间日期
    if condition.created_at_start:
        condition_array.append(BidPlagiarismCheckTask.created_at >= condition.created_at_start)
    if condition.created_at_end:
        condition_array.append(BidPlagiarismCheckTask.created_at <= condition.created_at_end)

    with app_context.db_session_factory() as session:
        tasks = session.query(BidPlagiarismCheckTask).filter(and_(*condition_array)).offset(offset).limit(per_page).all()
        count = session.query(BidPlagiarismCheckTask).filter(and_(*condition_array)).count()
        task_data = [{
            "id": task.id,
            "check_type": task.check_type,
            "task_name": task.task_name,
            "file_name_list": task.file_name_list,
            "process_status": task.process_status,
            "created_at": format_datetime(task.created_at)} for task in tasks]
    page = TenderTaskPage(
        page_offset=condition.page_offset,
        page_size=len(task_data),
        page_num=calculate_pages_int(count, per_page),
        total=count,
        data=task_data
    )
    return page


def get_tender_sub_task_list(task_id, base_page: BasePageDto):
    page = base_page.page_offset  # 当前页码（从 1 开始）
    per_page = base_page.page_size  # 每页记录数
    # 计算偏移量
    offset = (page - 1) * per_page
    with app_context.db_session_factory() as session:
        tasks = session.query(SubBidPlagiarismCheckTask).filter(SubBidPlagiarismCheckTask.bid_plagiarism_check_task_id == task_id).offset(offset).limit(per_page).all()
        count = session.query(SubBidPlagiarismCheckTask).filter(SubBidPlagiarismCheckTask.bid_plagiarism_check_task_id == task_id).count()
        task_data = [{
            "id": task.id,
            "left_file_name": task.left_file_name,
            "left_file_id": task.left_file_id,
            "right_file_id": task.right_file_id,
            "right_file_name": task.right_file_name,
            "similarity_number": task.similarity_number,
            "process_status": task.process_status} for task in tasks]
    tender_task_page = TenderTaskPage(
        page_offset=page,
        page_size=len(task_data),
        page_num=calculate_pages_int(count, per_page),
        total=count,
        data=task_data
    )
    return tender_task_page


def get_tender_similarity_info(sub_task_id)->TenderSimilarityVO:
    with app_context.db_session_factory() as session:
        sub_bid_plagiarism_check_task: SubBidPlagiarismCheckTask = session.get(SubBidPlagiarismCheckTask, sub_task_id)
        if not sub_bid_plagiarism_check_task:
            return None

    return get_tender_similarity_info_by_file_id(
        TenderSimilarityDto(sub_bid_plagiarism_check_task.left_file_id,
                            sub_bid_plagiarism_check_task.right_file_id))


# 定义表头结构
HEADERS = [
    "标书名称1",
    "页码",
    "相似片段",
    "标书名称2",
    "页码",
    "相似片段",
    "相似度"
]


def export_similarity_report_task_id(task_id):
    with app_context.db_session_factory() as session:
        sub_bid_plagiarism_check_tasks: SubBidPlagiarismCheckTask = session.query(SubBidPlagiarismCheckTask)\
            .filter(SubBidPlagiarismCheckTask.bid_plagiarism_check_task_id == task_id).all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "相似性对比"
        for sub_bid_plagiarism_check_task in sub_bid_plagiarism_check_tasks:
            export_similarity_report(sub_bid_plagiarism_check_task.id, wb)
        # 2. 创建内存流
        buffer = BytesIO()
        # 6. 保存 Excel 到内存流
        wb.save(buffer)
        buffer.seek(0)  # 将指针重置到开头
        return buffer


def export_similarity_report(sub_task_id, workbook: Workbook)-> Workbook:
    similarity_workbook(sub_task_id, workbook)


def similarity_workbook(sub_task_id, workbook: Workbook):
    with app_context.db_session_factory() as session:
        similarity_records = session.query(DocumentSimilarityRecord) \
            .filter(and_(DocumentSimilarityRecord.sub_bid_plagiarism_check_task_id == sub_task_id, DocumentSimilarityRecord.status == 1)) \
            .order_by(desc(DocumentSimilarityRecord.similarity)).all()
    if similarity_records:
        task_data = [{
            "left_file_name": record.left_file_name,
            "left_file_page": record.left_file_page,
            "left_file_page_chunk": record.left_file_page_chunk,
            "right_file_name": record.left_file_name,
            "right_file_page": record.right_file_page,
            "right_file_page_chunk": record.right_file_page_chunk,
            "similarity": record.similarity} for record in similarity_records]

        # 3. 初始化 Workbook
        wb = workbook
        ws = wb.active
        last_row_idx = ws.max_row
        if last_row_idx != 1:
            last_row_idx += 2
        # 4. 设置表头样式 (深蓝色背景，白色粗体字)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header_text in enumerate(HEADERS, 1):
            cell = ws.cell(row=last_row_idx, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # 设置默认列宽 (根据内容动态调整可放在写入数据后)
            ws.column_dimensions[chr(64 + col_idx)].width = 15

            # 5. 写入数据行 (隔行变色可选)
        # even_row_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        for row_idx, row_data in enumerate(task_data, last_row_idx+1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[value])

                # 偶数行填充浅色背景，提升可读性
                # if row_idx % 2 == 0:
                #     cell.fill = even_row_fill

                # 相似度列如果是高相似度，可以标红或标绿 (示例)
                if isinstance(value, str) and value.startswith('1'):
                    cell.font = Font(color="00FF00")  # 绿色表示高分

    return ws


def get_tender_similarity_info_by_file_id(tender_similarity_dto: TenderSimilarityDto)->TenderSimilarityVO:
    """
    根据两份标书的id查出相似度信息
    :param tender_similarity_dto: 查询条件
    {
        left_file_id: 左侧标书文件id
        right_file_id: 右侧标书文件id
    }
    :return:
    """
    page = tender_similarity_dto.page_offset  # 当前页码（从 1 开始）
    per_page = tender_similarity_dto.page_size  # 每页记录数
    # 计算偏移量
    offset = (page - 1) * per_page
    task_id = None
    task_data = []
    tender_reference_url = None
    with app_context.db_session_factory() as session:
        # 查询出相关的标书子任务
        sub_bid_plagiarism_check_task: SubBidPlagiarismCheckTask = session.query(SubBidPlagiarismCheckTask) \
            .filter(or_(and_(SubBidPlagiarismCheckTask.left_file_id == tender_similarity_dto.left_file_id,
                             SubBidPlagiarismCheckTask.right_file_id == tender_similarity_dto.right_file_id),
                        and_(SubBidPlagiarismCheckTask.left_file_id == tender_similarity_dto.right_file_id,
                             SubBidPlagiarismCheckTask.right_file_id == tender_similarity_dto.left_file_id))) \
            .first()
        # 是否存在两份标书的查重任务，如果不存在，直接结束，返回None
        if not sub_bid_plagiarism_check_task:
            return None
        # 标书任务的id
        task_id = sub_bid_plagiarism_check_task.bid_plagiarism_check_task_id
        # 查询标书的任务
        file_record: BidPlagiarismCheckTask = session.get(BidPlagiarismCheckTask, task_id)
        # 是否存在查重任务，如果不存在，直接结束，返回None
        if not file_record:
            return None
        # 获取招标文件id
        tender_reference_file_id = file_record.tender_reference_file_id
        # 获取任务所有的标书id "1,2,3,4"
        file_id_list: str = file_record.file_id_list
        # 获取查重任务中的所有标书信息
        if file_id_list:
            file_id_int_list = [int(x.strip()) for x in file_id_list.split(',')]
            file_record_entity_list = session.query(FileRecordEntity).filter(
                FileRecordEntity.id.in_(file_id_int_list)).all()
            file_record_vo: List = [
                FileRecordVO(id=file_record_item.id,
                             file_name=file_record_item.file_name,
                             file_url=get_file_url(file_record_item.file_path)) for file_record_item in file_record_entity_list]
        # 查询招标参考文件路径
        if tender_reference_file_id:
            file_record_entity: FileRecordEntity = session.get(FileRecordEntity, tender_reference_file_id)
            tender_reference_url = get_file_url(file_record_entity.file_path)
        if sub_bid_plagiarism_check_task:
            # 获取相似度数据列表，分页获取
            similarity_records = session.query(DocumentSimilarityRecord)\
                .filter(and_(DocumentSimilarityRecord.status == 1,
                            or_(and_(DocumentSimilarityRecord.left_file_id == tender_similarity_dto.left_file_id,
                                 DocumentSimilarityRecord.right_file_id == tender_similarity_dto.right_file_id),
                                and_(DocumentSimilarityRecord.left_file_id == tender_similarity_dto.right_file_id,
                                 DocumentSimilarityRecord.right_file_id == tender_similarity_dto.left_file_id))))\
                .order_by(desc(DocumentSimilarityRecord.similarity)).offset(offset).limit(per_page).all()

            # 获取列表总数
            count = session.query(DocumentSimilarityRecord) \
                .filter(and_(DocumentSimilarityRecord.status == 1,
                            or_(and_(DocumentSimilarityRecord.left_file_id == tender_similarity_dto.left_file_id,
                                 DocumentSimilarityRecord.right_file_id == tender_similarity_dto.right_file_id),
                            and_(DocumentSimilarityRecord.left_file_id == tender_similarity_dto.right_file_id,
                                 DocumentSimilarityRecord.right_file_id == tender_similarity_dto.left_file_id)))) \
                .count()
            task_data = [{
                "id": record.id,
                "left_file_name": record.left_file_name,
                "left_file_url": record.left_file_url,
                "left_file_page": record.left_file_page,
                "left_file_page_chunk": f"{record.left_file_page_chunk}({record.left_file_page})",
                "right_file_name": record.left_file_name,
                "right_file_url": record.right_file_url,
                "right_file_page": record.right_file_page,
                "right_file_page_chunk": f"{record.right_file_page_chunk}({record.right_file_page})",
                "similarity": record.similarity} for record in similarity_records]
    tender_similarity_vo = TenderSimilarityVO(
        page_offset=page,
        page_size=len(task_data),
        page_num=calculate_pages_int(count, per_page),
        total=count,
        data=task_data,
        tender_reference=tender_reference_url,
        tender_list=file_record_vo
    )
    return tender_similarity_vo


def update_tender_similarity_info(info_id):
    with app_context.db_session_factory() as session:
        record: DocumentSimilarityRecord = session.get(DocumentSimilarityRecord, info_id)
        record.status = 0
        session.add(record)
        session.commit()


def batch_update_tender_similarity_info(ids):
    with app_context.db_session_factory() as session:
        records: List[DocumentSimilarityRecord] = session.query(DocumentSimilarityRecord)\
            .filter(DocumentSimilarityRecord.id.in_(ids)).all()
        for record in records:
            record.status = 0
        session.add_all(records)
        session.commit()


def delete_tender_task_by_task_id(task_id: int):
    with app_context.db_session_factory() as session:
        task_record = session.get(BidPlagiarismCheckTask, task_id)
        session.query(SubBidPlagiarismCheckTask).filter_by(bid_plagiarism_check_task_id=task_id).delete()
        session.query(DocumentSimilarityRecord).filter_by(bid_plagiarism_check_task_id=task_id).delete()

        # 获取招标文件id
        # tender_reference_file_id = task_record.tender_reference_file_id
        # 获取任务所有的标书id "1,2,3,4"
        if task_record:
            file_id_list: str = task_record.file_id_list
            # 获取查重任务中的所有标书信息
            if file_id_list:
                file_id_int_list = [int(x.strip()) for x in file_id_list.split(',')]
                file_record_entity_list = session.query(FileRecordEntity).filter(
                    FileRecordEntity.id.in_(file_id_int_list)).all()
                for file_record_item in file_record_entity_list:
                    delete_object(file_record_item.file_path)
                    session.delete(file_record_item)
            session.delete(task_record)
            session.commit()


async def handle_tender_tender_reference_file(file_id: int):
    """
    处理标书招标文件
    :param file_id: 上传招标文件的id
    :return:
    """
    with app_context.db_session_factory() as session:
        file_record = session.get(FileRecordEntity, file_id)
        file_path = file_record.file_path
    vectorizer = QwenEmbeddingVectorizer()
    milvus_db = create_tender_reference_vector_milvus_db(vectorizer.get_vector_dim())
    business_id = file_record.business_id
    minio_client = app_context.minio_client
    with minio_client.get_object("tender", file_path) as response:
        file_data = response.read()  # 自动 close + release_conn
    pdf_stream = BytesIO(file_data)
    if file_record.mime_type == "pdf":
        pdf_parser = PdfParser()
        file_document = pdf_parser.parse_tender(stream=pdf_stream)
        documents: List[str] = pdf_parser.tender_overlapping_splitting(file_document, 100, 10)
        ems = vectorizer.encode(documents)
    file_ids = [file_id for document in documents]
    milvus_db.insert_info([file_ids, documents, ems])


def file_record_handle(file_record: TenderFile):
    """
    文件处理功能
    :param file_record: 文件管理数据表计入
    """
    print(f"处理标书：{file_record.file_name}")
    file_path = file_record.file_path
    business_id = file_record.business_id
    minio_client = app_context.minio_client
    with minio_client.get_object(business_id, file_path) as response:
        file_data = response.read()  # 自动 close + release_conn
    pdf_stream = BytesIO(file_data)
    if file_record.mime_type == "pdf":
        pdf_parser = PdfParser()
        if pdf_parser.is_scanned_pdf(stream=pdf_stream):
            print("扫描件")
            pdf_parser = MarkDownParser()

        print(f"解析标书：{file_record.file_name}")
        file_document = pdf_parser.parse(stream=pdf_stream, file_id=file_record.file_id)
        if file_document:
            embedding = QwenEmbeddingVectorizer()
            print(f"标书分片：{file_record.file_name}")
            documents: list[HDocument] = pdf_parser.overlapping_splitting(file_document, 100, 0)
            if len(documents) > 0:
                file_ids = []
                pages = []
                start_index_list = []
                texts = []
                for document in documents:
                    file_ids.append(document.file_id)
                    pages.append(document.page)
                    start_index_list.append(document.start_index)
                    texts.append(document.text)
                print(f"标书分片批量获取词向量：{file_record.file_name}")
                vec_list = embedding.encode_group(texts)
                print(f"标书向量数据生成：{file_record.file_name}")
                milvus_vector_db = create_tender_vector_milvus_db(embedding.get_vector_dim())
                print(f"标书批量入向量库：{file_record.file_name}")
                milvus_vector_db.insert_info([file_ids, pages, start_index_list, texts, vec_list])
    else:
        # 其他文件
        pass


def calculate_pages_int(total, per_page=10):
    """
    计算页数
    :param total:
    :param per_page:
    :return:
    """
    if total == 0:
        return 0
    return (total + per_page - 1) // per_page


class CheckTask:
    """
    检查标书任务
    """

    def __init__(self, file_record_a: TenderFile, file_record_b: TenderFile, tender_reference_id=None):
        self.file_record_a = file_record_a
        self.file_record_b = file_record_b
        self.tender_reference_id = tender_reference_id

    def rm_text(self, tender_item, milvus_reference_vector_db, rm_text_vector_milvus_db):
        # 获取所有的招标文件的向量数据信息
        if self.tender_reference_id:
            result_reference = milvus_reference_vector_db.search_similar(
                f"file_id == {self.tender_reference_id}", [tender_item["vector"]], None, 1)
            similarity_reference = []
            for reference_info in result_reference:
                if reference_info['similarity'] > 0.6:
                    similarity_reference.append(reference_info)
            if len(similarity_reference) > 0:
                return False
        rm_result = rm_text_vector_milvus_db.search_similar(f"", [tender_item["vector"]], None, 1)
        if rm_result and len(rm_result) > 0:
            similarity_rm = []
            for rm_item in rm_result:
                if rm_item['similarity'] > 0.8:
                    similarity_rm.append(rm_item)
            if len(similarity_rm) > 0:
                return False
        return True

    def execute(self):
        """
        执行比对任务
        """
        sub_id = None
        with app_context.db_session_factory() as session:
            # 创建查重任务
            sub_task = SubBidPlagiarismCheckTask(
                bid_plagiarism_check_task_id=self.file_record_a.tender_task_id,
                left_file_id=self.file_record_a.file_id,
                left_file_name=self.file_record_a.file_name,
                right_file_id=self.file_record_b.file_id,
                right_file_name=self.file_record_b.file_name,
                similarity_number=0
            )
            session.add(sub_task)
            session.commit()
            sub_id = sub_task.id
        file_url_a = get_file_url(self.file_record_a.file_path)
        file_url_b = get_file_url(self.file_record_b.file_path)
        milvus_topic_vector_db = create_tender_topic_vector_milvus_db(4096)
        milvus_vector_db = create_tender_vector_milvus_db(4096)
        milvus_reference_vector_db = create_tender_reference_vector_milvus_db(4096)
        rm_text_vector_milvus_db = create_rm_text_vector_milvus_db(4096)
        main_topic_vector_milvus_db = create_main_topic_vector_milvus_db(4096)
        # 获取标书a 的目录信息
        file_a_topic_list = milvus_topic_vector_db.query_data(f"tender_file_id == {self.file_record_a.file_id}",
                                                              ["topic_content", "vector"])
        document_similarity_records = []
        for file_a_topic in file_a_topic_list:
            logger.info(f"topic:{file_a_topic['topic_content']}")
            main_topic = main_topic_vector_milvus_db.search_similar(f"", [file_a_topic["vector"]], ["topic"], 1)
            if not main_topic or len(main_topic) < 1:
                logger.info(f"topic:{file_a_topic['topic_content']}; 不匹配正文")
                continue
            if main_topic[0]['similarity'] < 0.8:
                logger.info(f"topic:{file_a_topic['topic_content']}; 不匹配正文")
                continue
            logger.info(f"topic:{file_a_topic['topic_content']}; 匹配正文：{main_topic[0]['topic']}")

            # 匹配与b标书相似度最高的的章节
            result_topic_similarity = milvus_topic_vector_db.search_similar(f"tender_file_id == {self.file_record_b.file_id}",
                                                                            [file_a_topic["vector"]], ["topic_content"], 1)
            if not result_topic_similarity or len(result_topic_similarity) < 1:
                continue
            topic_content_similarity = result_topic_similarity[0].get('topic_content', None)
            if not topic_content_similarity:
                continue
            logger.info(f"标书-{self.file_record_a.file_id}，"
                        f"章节【{file_a_topic['topic_content']}】, "
                        f"关联标书-{self.file_record_b.file_id}: {topic_content_similarity}")
            topic_content = file_a_topic['topic_content']
            tender_a_data_list = milvus_vector_db.query_data(f"file_id == {self.file_record_a.file_id} and topic == '{topic_content}'",
                                                         ["file_id", "page", "start_index", "text_content", "vector"])
            for tender_item in tender_a_data_list:
                # 获取所有的招标文件的向量数据信息
                if not self.rm_text(tender_item, milvus_reference_vector_db, rm_text_vector_milvus_db):
                    continue
                result = milvus_vector_db.search_similar(f"file_id == {self.file_record_b.file_id} and topic == '{topic_content_similarity}'", [tender_item["vector"]])
                for info in result:
                    if info['similarity'] > 0.85:

                        document_similarity_record = DocumentSimilarityRecord(
                            bid_plagiarism_check_task_id=self.file_record_b.tender_task_id,
                            sub_bid_plagiarism_check_task_id=sub_id,
                            left_file_id=self.file_record_a.file_id,
                            left_file_name=self.file_record_a.file_name,
                            left_file_url=file_url_a,
                            left_file_page=tender_item['page'],
                            left_file_page_start_index=tender_item['start_index'],
                            left_file_page_chunk=tender_item['text_content'],
                            right_file_id=self.file_record_b.file_id,
                            right_file_name=self.file_record_b.file_name,
                            right_file_url=file_url_b,
                            right_file_page=info['page'],
                            right_file_page_start_index=info['start_index'],
                            right_file_page_chunk=info['text_content'],
                            similarity=info['similarity']
                        )
                        document_similarity_records.append(document_similarity_record)
        with app_context.db_session_factory() as session:
            session.add_all(document_similarity_records)
            sub_task: SubBidPlagiarismCheckTask = session.query(SubBidPlagiarismCheckTask).get(sub_id)
            sub_task.process_status = "completed"
            sub_task.similarity_number = len(document_similarity_records)
            session.add(sub_task)
            session.commit()
